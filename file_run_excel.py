import Data
import Function
import Neighborhood
import time
import copy
import Neighborhood11
import Neighborhood10
import Neighborhood_drone
import glob
import os
import openpyxl

global LOOP
global tabu_tenure
global best_sol
global best_fitness
global Tabu_Structure
global current_neighborhood
global LOOP_IMPROVED
global SET_LAST_10
global BEST

# Set up chỉ số -------------------------------------------------------------------
tabu_tenure = 18
tabu_tenure1 = 10
LOOP = 200
ITE = 2
epsilon = (-1) * 0.00001
BREAKLOOP = 40
LOOP_IMPROVED = 0
SET_LAST_10 = [] 
BEST = []


# Set up chỉ số -------------------------------------------------------------------

# Hai mẫu bị lỗi khi chạy bộ 10 C101 1 that vào ở initial solution để thử 
check = [[[[0, [1]], [1, [2, 5, 3, 4]], [2, [7]], [5, []], [3, []], [4, [10, 6, 8]], [10, []], [7, []], [6, [9]], [8, []], [9, []]]], [[[1, [2, 5, 3, 4]]], [[2, [7]]], [[4, [10, 6, 8]]], [[6, [9]]]]]
check2 = [[[[0, [1]], [1, [2, 5, 4]], [2, []], [5, [3, 10]], [3, [6]], [4, [7, 8]], [10, []], [7, []], [6, []], [8, [9]], [9, []]]], [[[1, [2, 5, 4]]], [[5, [3, 10]]], [[3, [6]]], [[4, [7, 8]]], [[8, [9]]]]]

check3 = [[[[0, [1]], [1, [2, 5, 3, 10]], [2, []], [5, [4, 6]], [3, [7]], [4, []], [10, []], [7, [8]], [6, [9]], [8, []], [9, []]]], [[[1, [2, 5, 3, 10]]], [[5, [4, 6]]], [[3, [7]]], [[7, [8]]], [[6, [9]]]]]
# Hai mẫu bị lỗi khi chạy bộ 10 C101 1 that vào ở initial solution để thử 
check6 = [[[[0, [1]], [5, [5, 2, 3, 4]], [4, []], [2, []], [1, []], [3, []], [7, [7, 6, 8, 10]], [10, []], [6, []], [8, []], [9, [9]]]], [[[5, [5, 2, 3, 4]]], [[7, [7, 6, 8, 10]]], [[9, [9]]]]]
# R101.3 15
check7 = [[[[0, [8]], [6, [6, 5]], [5, [7]], [8, [11, 10]], [7, []], [11, []], [10, [1]], [1, [3, 9]], [3, [12]], [9, []], [12, [4, 2]], [4, []], [2, []], [13, [13, 15, 14]], [15, []], [14, []]]], [[[6, [5, 6]]], [[5, [7]]], [[8, [11, 10]]], [[10, [1]]], [[1, [3, 9]]], [[3, [12]]], [[12, [2, 4]]], [[13, [13, 14, 15]]]]]


def Tabu_search_for_CVRP():
    # print(Data.standard_deviation)
    global current_neighborhood
    global LOOP_IMPROVED
    
    Tabu_Structure = [tabu_tenure * (-1)] * Data.number_of_cities
    Tabu_Structure1 = []
    for i in range(Data.number_of_cities):
        row = [tabu_tenure * (-1)] * Data.number_of_cities
        Tabu_Structure1.append(row)
        
    Tabu_Structure2 = []
    for i in range(Data.number_of_cities):
        row = [tabu_tenure * (-1)] * Data.number_of_cities
        Tabu_Structure2.append(row)
        
    list_init = []
    current_sol1 = Function.initial_solution1()
    current_sol2 = Function.initial_solution()
    current_sol3 = Function.initial_solution3()
    current_sol4 = Function.initial_solution4()
    # current_sol5 = Function.initial_solution5()
    
    current_sol1 = Neighborhood.Optimize_initial_solution_in_drone(current_sol1)
    current_sol2 = Neighborhood.Optimize_initial_solution_in_drone(current_sol2)
    current_sol3 = Neighborhood.Optimize_initial_solution_in_drone(current_sol3)
    current_sol4 = Neighborhood.Optimize_initial_solution_in_drone(current_sol4)
    
    list_init.append(current_sol1)
    list_init.append(current_sol2)
    list_init.append(current_sol3)
    list_init.append(current_sol4)
    # list_init.append(current_sol5)

    
    
    list_fitness_init = []
    fitness1 = Function.fitness(current_sol1)
    fitness2 = Function.fitness(current_sol2)
    fitness3 = Function.fitness(current_sol3)
    fitness4 = Function.fitness(current_sol4)
    # fitness5 = Function.fitness(current_sol5)

    list_fitness_init.append(fitness1)
    list_fitness_init.append(fitness2)
    list_fitness_init.append(fitness3)
    list_fitness_init.append(fitness4)
    # list_fitness_init.append(fitness5)

    
    current_fitness = list_fitness_init[0][0]
    current_sol = list_init[0]
    # for i in range(len(list_init)):
    #     print(list_init[i])
    #     print(list_fitness_init[i])
    #     print("--------")
    for i in range(1, len(list_fitness_init)):
        # print("init-----",i)
        # print(list_init[i])
        # print(list_fitness_init[i][0])
        if current_fitness > list_fitness_init[i][0]:
            current_sol = list_init[i]
            current_fitness = list_fitness_init[i][0]

    # Initial solution thay ở đây ------------->
    # current_sol = check7     # Để dòng này làm comment để tìm initial solution theo tham lam
    # <------------- Initial solution thay ở đây 
    
    current_fitness, current_truck_time = Function.fitness(current_sol)
    best_sol = current_sol
    best_fitness = current_fitness
    sol_chosen_to_break = current_sol
    fit_of_sol_chosen_to_break = current_fitness
    print(best_sol) 
    print(best_fitness)
    print(Function.Check_if_feasible(best_sol))
    for i in range(LOOP):
        # print("------------------",i,"------------------")
        current_neighborhood = []
        if i - LOOP_IMPROVED > BREAKLOOP:
            # current_neighborhood5 = Neighborhood.swap_two_array(sol_chosen_to_break)
            # print(sol_chosen_to_break)
            current_neighborhood9 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.swap_two_array, sol_chosen_to_break, 3, 2, False)
            BEST.append(sol_chosen_to_break)
            current_neighborhood.append([9, current_neighborhood9])
            fit_of_sol_chosen_to_break = 10000000 
            
            for j in range(len(Tabu_Structure)):
                Tabu_Structure[j] -= tabu_tenure
                
        elif i - LOOP_IMPROVED == BREAKLOOP - 3:
            new_solution, new_fitness, if_improved = Neighborhood.find_if_truck_route_need_reverse(sol_chosen_to_break)
            if if_improved:
                current_sol = new_solution
                current_fitness = new_fitness
                current_truck_time = Function.fitness(new_solution)[1]
                LOOP_IMPROVED = i
                for j in range(len(Tabu_Structure)):
                    Tabu_Structure[j] -= tabu_tenure
                for i in range(len(Tabu_Structure1)):
                    for j in range(len(Tabu_Structure1[i])):
                        Tabu_Structure1[i][j] -= tabu_tenure1
                
                if new_fitness < best_fitness:
                    best_sol = new_solution
                    best_fitness = new_fitness
            continue
        elif i - LOOP_IMPROVED == int(BREAKLOOP/2):
            current_neighborhood10 = Neighborhood11.Neighborhood_two_opt(sol_chosen_to_break)
            current_neighborhood.append([10, current_neighborhood10])
        elif i % 40 == 15:
            # current_neighborhood6point5 = Neighborhood.one_opt_and_change_truck_route_after(current_sol, current_truck_time)
            # current_neighborhood6 = Neighborhood.one_opt_and_change_truck_route_after(current_sol, current_truck_time)
            current_neighborhood7point5 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.one_opt_and_change_truck_route_after, sol_chosen_to_break, 1, 2, True)
            current_neighborhood7 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.one_opt_and_change_truck_route_after, current_sol, 2, 2, True)
            current_neighborhood7 = current_neighborhood7 + current_neighborhood7point5
            current_neighborhood.append([7, current_neighborhood6])
            for j in range(len(Tabu_Structure)):
                Tabu_Structure[j] = i - ( tabu_tenure - 2)
            if i - LOOP_IMPROVED > BREAKLOOP - tabu_tenure:
                LOOP_IMPROVED = i - (BREAKLOOP - tabu_tenure)
        elif i % 6 == 3:
            current_neighborhood1 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list(name_of_truck_neiborhood=Neighborhood10.Neighborhood_one_otp, solution=current_sol, number_of_potial_solution=3, number_of_loop_drone=2, tabu_list=Tabu_Structure, tabu_tenure=tabu_tenure,  index_of_loop=i, best_fitness=best_fitness, is_1_0=True)
            current_neighborhood.append([1, current_neighborhood1])
        elif i % 12 == 3:
            current_neighborhood3 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list(name_of_truck_neiborhood=Neighborhood11.Neighborhood_move_1_1_ver2, solution=current_sol, number_of_potial_solution=3, number_of_loop_drone=2, tabu_list=Tabu_Structure1, tabu_tenure=tabu_tenure1,  index_of_loop=i, best_fitness=best_fitness, is_1_0=False)
            current_neighborhood.append([3, current_neighborhood3])
        elif i % 6 < 4:
            # current_neighborhood1 = Neighborhood.Neighborhood_one_otp_fix(current_sol)
            current_neighborhood1 = Neighborhood10.Neighborhood_one_otp(current_sol, current_truck_time)
            current_neighborhood2 = Neighborhood10.Neighborhood_one_otp_plus(current_sol, current_truck_time)
            current_neighborhood3 = Neighborhood11.Neighborhood_move_1_1_ver2(current_sol)
            # current_neighborhood3 = Neighborhood11.two_swap(current_sol, current_truck_time)
            # current_neighborhood4 = Neighborhood11.Neighborhood_two_opt(current_sol)
            
            current_neighborhood.append([1, current_neighborhood1])
            current_neighborhood.append([2, current_neighborhood2])
            current_neighborhood.append([3, current_neighborhood3])
            # current_neighborhood.append([4, current_neighborhood4])
        # elif i % 7 < 5:
        #     current_neighborhood3 = Neighborhood11.Neighborhood_move_1_1_ver2(current_sol)
        #     current_neighborhood3 = Neighborhood11.two_swap(current_sol, current_truck_time)
        #     current_neighborhood4 = Neighborhood11.Neighborhood_two_opt(current_sol)
        #     current_neighborhood.append([3, current_neighborhood3])
        #     current_neighborhood.append([4, current_neighborhood4])
        else:
            # current_neighborhood5 = Neighborhood_drone.Neighborhood_group_trip(current_sol)
            current_neighborhood6 = Neighborhood_drone.Neighborghood_change_drone_route_max_pro_plus(current_sol)
        
            # current_neighborhood.append([5, current_neighborhood5])
            current_neighborhood.append([6, current_neighborhood6])
        index = [-1] * len(current_neighborhood)
        min_nei = [100000] * len(current_neighborhood)
        for j in range(len(current_neighborhood)):
            if current_neighborhood[j][0] in [1, 2]:
                for k in range(len(current_neighborhood[j][1])):
                    cfnode = current_neighborhood[j][1][k][1][0]
                    if cfnode - best_fitness < epsilon:
                        min_nei[j] = cfnode
                        index[j] = k
                        best_fitness = cfnode
                        best_sol = current_neighborhood[j][1][k][0]
                        LOOP_IMPROVED = i

                    elif cfnode - min_nei[j] < epsilon and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= i:
                        min_nei[j] = cfnode
                        index[j] = k
            elif current_neighborhood[j][0] == 3:
                for k in range(len(current_neighborhood[j][1])):    
                    cfnode = current_neighborhood[j][1][k][1][0]
                    if cfnode - best_fitness < epsilon:
                        min_nei[j] = cfnode
                        index[j] = k
                        best_fitness = cfnode
                        best_sol = current_neighborhood[j][1][k][0]
                        LOOP_IMPROVED = i

                    elif cfnode - min_nei[j] < epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]][current_neighborhood[j][1][k][2][1]] + tabu_tenure <= i:
                        min_nei[j] = cfnode
                        index[j] = k
            elif current_neighborhood[j][0] == 4:
                for k in range(len(current_neighborhood[j][1])):
                    cfnode = current_neighborhood[j][1][k][1][0]
                    if cfnode - best_fitness < epsilon:
                        min_nei[j] = cfnode
                        index[j] = k
                        best_fitness = cfnode
                        best_sol = current_neighborhood[j][1][k][0]
                        LOOP_IMPROVED = i

                    elif cfnode - min_nei[j] < epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]][current_neighborhood[j][1][k][2][1]] + tabu_tenure <= i:
                        min_nei[j] = cfnode
                        index[j] = k
            else:
                for k in range(len(current_neighborhood[j][1])):
                    cfnode = current_neighborhood[j][1][k][1][0]
                    if cfnode - best_fitness < epsilon:
                        min_nei[j] = cfnode
                        index[j] = k
                        best_fitness = cfnode
                        best_sol = current_neighborhood[j][1][k][0]
                        LOOP_IMPROVED = i
                        
                    elif cfnode - min_nei[j] < epsilon:
                        min_nei[j] = cfnode
                        index[j] = k
        index_best_nei = 0
        best_fit_in_cur_loop = min_nei[0]
        for j in range(1, len(min_nei)):
            if min_nei[j] < best_fit_in_cur_loop:
                index_best_nei = j
                best_fit_in_cur_loop = min_nei[j]
        if len(current_neighborhood[index_best_nei][1]) == 0:
            continue
        if current_neighborhood[index_best_nei][0] == 9:
            LOOP_IMPROVED = i
        current_sol = current_neighborhood[index_best_nei][1][index[index_best_nei]][0]
        current_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][0]
        current_truck_time = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][1]
        
        # SET_LAST_10.append([current_sol, [current_fitness, current_truck_time]])
        # if len(SET_LAST_10) > 10:
        #     SET_LAST_10.pop(0)
            
        if current_neighborhood[index_best_nei][0] in [1, 2]:
            Tabu_Structure[current_neighborhood[index_best_nei][1][index[index_best_nei]][2]] = i
        
        if current_neighborhood[index_best_nei][0] == 3:
            Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]][current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = i
            Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]][current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = i
        
        if current_neighborhood[index_best_nei][0] == 4:
            Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]][current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = i
            Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]][current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = i

        if fit_of_sol_chosen_to_break > current_fitness:
            sol_chosen_to_break = current_sol
            fit_of_sol_chosen_to_break = current_fitness
        # if current_neighborhood[index_best_nei][0] == 6:
        #     print("------------------",i,"------------------")
        #     print(current_neighborhood[index_best_nei][0])
        #     print(current_sol)
        #     print(current_fitness)
        
        # print("------------------",i,"------------------")
        # print(current_neighborhood[index_best_nei][0])
        # print(current_sol)
        # print(current_fitness)
        # print(Function.Check_if_feasible(current_sol))
    print("--before post optimization--")
    print(best_fitness)
    
    for ii in range(len(BEST)):
        
        option = [Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood, Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood, Neighborhood10.Neighborhood_move_depot
                ]
        
        bet_fit, best_truck_time = Function.fitness(BEST[ii])
        list_neighborhood_change_truck_route = [Neighborhood10.Neighborhood_one_otp, Neighborhood10.Neighborhood_one_otp_plus]
        for i in range(len(option)):
            stop = True
            while stop:
                # print(i)
                stop = False
                if i == 0:
                    neighborhood = option[i](list_neighborhood_change_truck_route[0], BEST[ii], 15, 2, True)
                elif i == 1:
                    neighborhood = option[i](list_neighborhood_change_truck_route[1], BEST[ii], 15, 2, True)
                elif i == 2:
                    neighborhood = option[i](BEST[ii])
                for j in range(len(neighborhood)):
                    cfnode = neighborhood[j][1][0]
                    if cfnode - bet_fit < epsilon:
                        # print(i,"----",cfnode)
                        bet_fit = cfnode
                        BEST[ii] = neighborhood[j][0]
                        best_truck_time = neighborhood[j][1][1]
                        stop = True
        if bet_fit - best_fitness < epsilon:
            best_sol = BEST[ii]
            best_fitness = bet_fit
          
    return best_fitness, best_sol


# Thư mục chứa các file .txt
folder_path = "test_data\\Smith\\TSPrd(time)\\Solomon\\15\\"

# Danh sách tất cả các file .txt trong thư mục
txt_files = glob.glob(os.path.join(folder_path, "*.dat"))

# Tạo một tệp Excel mới
workbook = openpyxl.Workbook()
sheet = workbook.active

# Dòng và cột bắt đầu ghi kết quả
row = 1
column = 1

# Ghi tên file .txt vào cột đầu tiên
for txt_file in txt_files:
    sheet.cell(row=row, column=column, value=os.path.basename(txt_file))
    row += 1
# Đặt lại dòng và cột cho việc ghi kết quả
row = 1
for txt_file in txt_files:
    column = 2
    with open(txt_file, 'r') as file:
        # Đọc nội dung từ file .txt và xử lý nó
        # print(txt_file)
        Data.read_data(txt_file)
        result = []
        run_time = []
        avg = 0
        # Tính kết quả từ nội dung file .txt (điều này phụ thuộc vào logic của chương trình của bạn)
        for i in range(ITE):
            BEST = []
            print("------------------------",i,"------------------------")
            start_time = time.time()
            best_fitness, best_sol = Tabu_search_for_CVRP()
            print("---------- RESULT ----------")
            print(best_sol)
            print(best_fitness)
            avg += best_fitness/ITE
            result.append(best_fitness)
            print(Function.Check_if_feasible(best_sol))
            end_time = time.time()
            run = end_time - start_time
            run_time.append(run)
            sheet.cell(row=row, column=column, value=best_fitness)
            column += 1
            workbook.save("Book1.xlsx")
        # Tăng dòng cho lần chạy tiếp theo
        row += 1

# Lưu kết quả vào tệp Excel
# workbook.save("Book2.xlsx")

# Đóng tệp Excel
workbook.close()
