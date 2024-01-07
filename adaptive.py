import Data
import Function
import Neighborhood
import time
import random
import csv
import Neighborhood11
import Neighborhood10
import Neighborhood_drone
import numpy as np
import openpyxl
import copy

global LOOP
global tabu_tenure
global best_sol
global best_fitness
global Tabu_Structure
global current_neighborhood
global LOOP_IMPROVED
global SET_LAST_10
global BEST

# Set up chỉ số -------------------------------------------------------------------
# 15:   50 - 35,    20:    60 - 40    10:   20 - 15
tabu_tenure = 10
tabu_tenure1 = tabu_tenure3 = tabu_tenure2 = 7
LOOP = 60
ITE = 3
SEGMENT = 5
epsilon = (-1) * 0.00001
# 15:   120,    20:    150
BREAKLOOP = 10
LOOP_IMPROVED = 0
SET_LAST_10 = [] 
BEST = []
time_limit = 800
time_limit_seg = 200
time_limit_post_opt = 100

# Set up chỉ số -------------------------------------------------------------------

# Hai mẫu bị lỗi khi chạy bộ 10 C101 1 that vào ở initial solution để thử 
check = [[[[0, [1]], [1, []], [8, [8, 14, 15]], [5, [5, 2, 13, 6]], [14, []], [15, []], [2, []], [13, []], [6, []]], [[0, []], [12, [12, 4, 3, 10]], [4, []], [3, []], [9, [9, 11, 7]], [10, []], [11, []], [7, []]]], [[[8, [8, 14, 15]]], [[5, [5, 2, 13, 6]]], [[12, [12, 4, 3, 10]]], [[9, [9, 11, 7]]]]]

def roulette_wheel_selection(population, fitness_scores):
    total_fitness = sum(fitness_scores)
    probabilities = [score / total_fitness for score in fitness_scores]
    selected_index = np.random.choice(len(population), p=probabilities)
    return population[selected_index]

def Tabu_search_for_CVRP():
    # print(Data.standard_deviation)
    global current_neighborhood
    global LOOP_IMPROVED
    
    Data1 = [['act', 'fitness', 'change1', 'change2', 'solution', 'tabu structue', 'tabu structure1']]
    
    Tabu_Structure = [tabu_tenure * (-1)] * Data.number_of_cities
    Tabu_Structure1 = [tabu_tenure1 * (-1)] * Data.number_of_cities
    Tabu_Structure2 = [tabu_tenure1 * (-1)] * Data.number_of_cities
    Tabu_Structure3 = [tabu_tenure1 * (-1)] * Data.number_of_cities
        
    # Tabu_Structure2 = []
    # for i in range(Data.number_of_cities):
    #     row = [tabu_tenure * (-1)] * Data.number_of_cities
    #     Tabu_Structure2.append(row)
    st_time = time.time()
    list_init = []
    #current_sol1 = Function.initial_solution1()
    #current_sol2 = Function.initial_solution()
    #current_sol3 = Function.initial_solution3()
    #current_sol4 = Function.initial_solution4()
    current_sol5 = Function.initial_solution5()
    
    #current_sol1 = Neighborhood.Optimize_initial_solution_in_drone(current_sol1)
    #current_sol2 = Neighborhood.Optimize_initial_solution_in_drone(current_sol2)
    #current_sol3 = Neighborhood.Optimize_initial_solution_in_drone(current_sol3)
    #current_sol4 = Neighborhood.Optimize_initial_solution_in_drone(current_sol4)
    
    #list_init.append(current_sol1)
    #list_init.append(current_sol2)
    #list_init.append(current_sol3)
    #list_init.append(current_sol4)
    list_init.append(current_sol5)

    
    
    list_fitness_init = []
    #fitness1 = Function.fitness(current_sol1)
    #fitness2 = Function.fitness(current_sol2)
    #fitness3 = Function.fitness(current_sol3)
    #fitness4 = Function.fitness(current_sol4)
    fitness5 = Function.fitness(current_sol5)

    #list_fitness_init.append(fitness1)
    #list_fitness_init.append(fitness2)
    #list_fitness_init.append(fitness3)
    #list_fitness_init.append(fitness4)
    list_fitness_init.append(fitness5)

    
    current_fitness = list_fitness_init[0][0]
    current_sol = list_init[0]
    # for i in range(len(list_init)):
    #     print(list_init[i])
    #     print(list_fitness_init[i])
    #     print("--------")
    for i in range(1, len(list_fitness_init)):
        if current_fitness > list_fitness_init[i][0]:
            current_sol = list_init[i]
            current_fitness = list_fitness_init[i][0]

    # Initial solution thay ở đây ------------->
    # current_sol = check     # Để dòng này làm comment để tìm initial solution theo tham lam
    # <------------- Initial solution thay ở đây 
    
    current_fitness, current_truck_time, current_sum_fitness = Function.fitness(current_sol)
    best_sol = current_sol
    best_fitness = current_fitness
    sol_chosen_to_break = current_sol
    fit_of_sol_chosen_to_break = current_fitness
    print(best_sol) 
    print(best_fitness)
    print(Function.Check_if_feasible(best_sol))
    T = 0
    nei_set = [0, 1, 2, 3, 4]
    weight = [1/len(nei_set)]*len(nei_set)
    step = [0, 0]

    while(True):
        #factor = random.uniform(0.1, 0.3)
        factor = 0.5 #0.3 0.6
        score = [0]*len(nei_set)
        used = [0]*len(nei_set)
        st = time.time()
        for i in range(LOOP):
            prev_fitness = current_fitness
            choose = roulette_wheel_selection(nei_set, weight)
            # print("------------------",i,"------------------")
            current_neighborhood = []
            if i - LOOP_IMPROVED > BREAKLOOP:
                print('break1')
                # current_neighborhood5 = Neighborhood.swap_two_array(sol_chosen_to_break)
                # print(sol_chosen_to_break)
                current_neighborhood10 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.swap_two_array, sol_chosen_to_break, 5, 3, False)
                BEST.append(sol_chosen_to_break)
                current_neighborhood.append([10, current_neighborhood10])
                fit_of_sol_chosen_to_break = 10000000 
                
                for j in range(len(Tabu_Structure)):
                    Tabu_Structure[j] -= tabu_tenure
                    Tabu_Structure1[j] -= tabu_tenure1
                    Tabu_Structure2[j] -= tabu_tenure2
                    Tabu_Structure3[j] -= tabu_tenure3
                
            elif i - LOOP_IMPROVED == BREAKLOOP - 3:
                print('break2')
                new_solution, new_fitness, if_improved = Neighborhood.find_if_truck_route_need_reverse(sol_chosen_to_break)
                if if_improved:
                    current_sol = new_solution
                    current_fitness = new_fitness
                    current_truck_time = Function.fitness(new_solution)[1]
                    LOOP_IMPROVED = i
                    for j in range(len(Tabu_Structure)):
                        Tabu_Structure[j] -= tabu_tenure
                        Tabu_Structure1[j] -= tabu_tenure1
                        Tabu_Structure2[j] -= tabu_tenure2
                        Tabu_Structure3[j] -= tabu_tenure3
                    if new_fitness < best_fitness:
                        best_sol = new_solution
                        best_fitness = new_fitness
                continue
            # elif i - LOOP_IMPROVED == int(BREAKLOOP/2):
            #     current_neighborhood10 = Neighborhood11.Neighborhood_two_opt(sol_chosen_to_break)
            #     current_neighborhood.append([10, current_neighborhood10])
            #elif i % 30 == 15:
            # elif i - LOOP_IMPROVED == BREAKLOOP - 50:
                # current_neighborhood6point5 = Neighborhood.one_opt_and_change_truck_route_after(current_sol, current_truck_time)
                # current_neighborhood6 = Neighborhood.one_opt_and_change_truck_route_after(current_sol, current_truck_time)
                
                # for j in range(len(Tabu_Structure)):
                #     Tabu_Structure[j] = i - ( tabu_tenure - 5)
                #     Tabu_Structure1[j] = i - ( tabu_tenure1 - 5)
                # for j in range(len(Tabu_Structure)):
                #     Tabu_Structure[j] -= 10
                #     Tabu_Structure1[j] -= 10
                # if i - LOOP_IMPROVED > BREAKLOOP - tabu_tenure:
                #     LOOP_IMPROVED = i - (BREAKLOOP - tabu_tenure)
       
                # for j in range(len(Tabu_Structure)):
                #         Tabu_Structure[j] -= tabu_tenure
                #         Tabu_Structure1[j] -= tabu_tenure1
                #         Tabu_Structure2[j] -= tabu_tenure2
                #         Tabu_Structure3[j] -= tabu_tenure3
            elif i % 10 < 7:
                # current_neighborhood1 = Neighborhood.Neighborhood_one_otp_fix(current_sol)
                if choose == 0:
                    current_neighborhood4 = Neighborhood11.Neighborhood_move_2_1(current_sol)
                    current_neighborhood5 = Neighborhood11.Neighborhood_two_opt_tue(current_sol)
                    current_neighborhood.append([5, current_neighborhood5])
                    current_neighborhood.append([4, current_neighborhood4])
                elif choose == 1:
                    current_neighborhood3 = Neighborhood11.Neighborhood_move_1_1_ver2(current_sol)
                    current_neighborhood.append([3, current_neighborhood3])
                elif choose == 2:
                    current_neighborhood1 = Neighborhood10.Neighborhood_one_otp(current_sol, current_truck_time)
                    current_neighborhood.append([1, current_neighborhood1])
                elif choose == 3:
                    current_neighborhood9point5 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.one_opt_and_change_truck_route_after, sol_chosen_to_break, 5, 3, True)
                    current_neighborhood9 = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood(Neighborhood.one_opt_and_change_truck_route_after, current_sol, 5, 3, True)
                    current_neighborhood9 = current_neighborhood9 + current_neighborhood9point5
                    current_neighborhood.append([9, current_neighborhood9])
                    if i - LOOP_IMPROVED > BREAKLOOP - 7:
                        LOOP_IMPROVED += 10
                else:
                    current_neighborhood2 = Neighborhood10.Neighborhood_one_otp_plus(current_sol, current_truck_time)
                    current_neighborhood.append([2, current_neighborhood2])
                
                # current_neighborhood4 = Neighborhood11.Neighborhood_move_2_1(current_sol)
                # current_neighborhood5 = Neighborhood11.Neighborhood_two_opt_tue(current_sol)
                # current_neighborhood.append([4, current_neighborhood4])
                # current_neighborhood.append([5, current_neighborhood5])
                # current_neighborhood3 = Neighborhood11.Neighborhood_move_1_1_ver2(current_sol)
                # current_neighborhood.append([3, current_neighborhood3])
                # current_neighborhood1 = Neighborhood10.Neighborhood_one_otp(current_sol, current_truck_time)
                # current_neighborhood2 = Neighborhood10.Neighborhood_one_otp_plus(current_sol, current_truck_time)
                # current_neighborhood.append([2, current_neighborhood2])
                # current_neighborhood.append([1, current_neighborhood1])
            else:
                # current_neighborhood7 = Neighborhood_drone.Neighborhood_group_trip(current_sol)
                # current_neighborhood.append([7, current_neighborhood7])
                current_neighborhood6 = Neighborhood_drone.Neighborghood_change_drone_route_max_pro_plus(current_sol)
                current_neighborhood.append([6, current_neighborhood6])
                current_neighborhood8 = Neighborhood_drone.Neighborhood_change_index_trip(current_sol)
                current_neighborhood.append([8, current_neighborhood8])
                # a = random.random()
                # if a < 0.7:
                #     current_neighborhood6 = Neighborhood_drone.Neighborghood_change_drone_route_max_pro_plus(current_sol)
                #     current_neighborhood.append([6, current_neighborhood6])
                # else:
                #     current_neighborhood11 = Neighborhood_drone.Neighborhood_change_index_trip(current_sol)
                #     current_neighborhood.append([11, current_neighborhood11])
            index = [-1] * len(current_neighborhood)
            min_nei = [100000] * len(current_neighborhood)
            min_sum = [1000000000] * len(current_neighborhood)
            flag = False
            for j in range(len(current_neighborhood)):
                if current_neighborhood[j][0] in [1, 2]:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= i:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]
                        elif min_nei[j] - epsilon > cfnode and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= i:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 3:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= i or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= i:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = min_sum[j] = current_neighborhood[j][1][k][1][2]
                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= i or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= i:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 4:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= i or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= i or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= i:
                            min_nei[j] = cfnode
                            index[j] = k
                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= i or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= i or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= i:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 5:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= i or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= i:
                            min_nei[j] = cfnode
                            index[j] = k
                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= i or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= i:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                else:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True
                            
                        elif cfnode - min_nei[j] < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                        elif cfnode < min_nei[j] - epsilon:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
            index_best_nei = 0
            best_fit_in_cur_loop = min_nei[0]
            
            # for j in range(len(min_nei)):
            #     print(min_nei[j])
            #     print(current_neighborhood[j][1][index[j]][0])
            #     print("-------")
            
            for j in range(1, len(min_nei)):
                if min_nei[j] < best_fit_in_cur_loop:
                    index_best_nei = j
                    best_fit_in_cur_loop = min_nei[j]
            if len(current_neighborhood[index_best_nei][1]) == 0:
                continue
            if current_neighborhood[index_best_nei][0] == 10:
                LOOP_IMPROVED = i
            current_sol = current_neighborhood[index_best_nei][1][index[index_best_nei]][0]
            current_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][0]
            current_truck_time = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][1]
            current_sum_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][2]
            # SET_LAST_10.append([current_sol, [current_fitness, current_truck_time]])
            # if len(SET_LAST_10) > 10:
            #     SET_LAST_10.pop(0)
                
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                Tabu_Structure[current_neighborhood[index_best_nei][1][index[index_best_nei]][2]] = i
            
            if current_neighborhood[index_best_nei][0] == 3:
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = i
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = i
            
            if current_neighborhood[index_best_nei][0] == 4:
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = i
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = i
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][2]] = i

            if current_neighborhood[index_best_nei][0] == 5:
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = i
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = i
                
            if fit_of_sol_chosen_to_break > current_fitness:
                sol_chosen_to_break = current_sol
                fit_of_sol_chosen_to_break = current_fitness
                LOOP_IMPROVED = i
            # if current_neighborhood[index_best_nei][0] == 6:
            #     print("------------------",i,"------------------")
            #     print(current_neighborhood[index_best_nei][0])
            #     print(current_sol)
            #     print(current_fitness)
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2], -1, current_sol, Tabu_Structure, Tabu_Structure1]
            elif current_neighborhood[index_best_nei][0] in [3]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0], current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1], current_sol, Tabu_Structure, Tabu_Structure1]
            else:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, -1, -1, current_sol]
            Data1.append(temp)
            if i % 10 < 7:
                used[choose] += 1
                if flag == True:
                    score[choose] += 0.5
                elif current_fitness - prev_fitness < epsilon:
                    score[choose] += 0.3
                else:
                    score[choose] += 0.1
                for j in range(len(nei_set)):
                    if used[j] == 0:
                        continue
                    else:
                        weight[j] = (1 - factor)*weight[j] + factor*score[j]/used[j]
            if flag == True:
                step = [LOOP_IMPROVED, T]
            ed = time.time()
            if ed - st > time_limit_seg: break
            cur_time = time.time()
            elapsed_time = cur_time - st_time
            if elapsed_time > time_limit: break
            print("------------------",i,"------------------")
            print(current_neighborhood[index_best_nei][0])
            print(current_sol)
            print(current_fitness)
            print(Function.Check_if_feasible(current_sol))
            print(Function.cal_truck_time(current_sol))
            print(Tabu_Structure)
            print(Tabu_Structure1)
            print(Tabu_Structure2)
            print(Tabu_Structure3)
       
        print(elapsed_time)
        print("-------T--------")
        print(best_fitness)
        print(T, best_sol, "\n", best_fitness)
        print(used, score)
        T +=1
        if elapsed_time > time_limit: break
        

    s = time.time()
    for ii in range(len(BEST)):
        
        option = [Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood, Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood,
                Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood, Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood,
                Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood, Neighborhood10.Neighborhood_move_depot
                ]
        
        bet_fit, best_truck_time, best_sum = Function.fitness(BEST[ii])
        list_neighborhood_change_truck_route = [Neighborhood10.Neighborhood_one_otp, Neighborhood10.Neighborhood_one_otp_plus, 
                                                Neighborhood11.Neighborhood_move_1_1_ver2, Neighborhood11.Neighborhood_move_2_1,
                                                Neighborhood11.Neighborhood_two_opt_tue]
        for i in range(len(option)):
            stop = True
            while stop:
                # print(i)
                stop = False
                if i in [0, 1]:
                    neighborhood = option[i](list_neighborhood_change_truck_route[i], BEST[ii], 15, 2, True)
                elif i in [2, 3, 4]:
                    neighborhood = option[i](list_neighborhood_change_truck_route[i], BEST[ii], 15, 2, False)
                else:
                    neighborhood = option[i](BEST[ii])
                for j in range(len(neighborhood)):
                    cfnode = neighborhood[j][1][0]
                    if cfnode - bet_fit < epsilon:
                        # print(i,"----",cfnode)
                        bet_fit = cfnode
                        BEST[ii] = neighborhood[j][0]
                        best_truck_time = neighborhood[j][1][1]
                        stop = True
        if bet_fit - best_fitness < epsilon:
            best_sol = BEST[ii]
            best_fitness = bet_fit

        c = time.time()
        if c - s > time_limit_post_opt:
            break  

    return best_fitness, best_sol, step



dataList = []
datatype = ['C101_', 'C201_', 'R101_', 'RC101_']
beta = [0.5, 1, 1.5, 2, 2.5, 3]
for i in range(len(datatype)):
    for j in range(len(beta)):
        datatemp = str(datatype[i]) + str(beta[j]) + '.dat'
        dataList.append(datatemp)


for k in range(len(dataList)):
    print(k, dataList[k])
    to_run = str(Data.file_path) + "\\" + str(dataList[k])
    Data.read_data_random(to_run)
    # to_run = str(Data.file_path) + "\\" + "C101_1.dat"
    # Data.read_data(to_run)

    result = []
    step_avg = []
    step_avg1 = []
    run_time = []
    avg = 0
    sol = []
    prev_fitness = 999999999999 
    for i in range(ITE):
        BEST = []
        print("------------------------",i,"------------------------")
        start_time = time.time()
        best_fitness, best_sol, step = Tabu_search_for_CVRP()
        print("---------- RESULT ----------")
        print(best_sol)
        print(best_fitness)
        avg += best_fitness/ITE
        result.append(best_fitness)
        step_avg.append(step[0])
        step_avg1.append(step[1])
        #print(Function.Check_if_feasible(best_sol))
        end_time = time.time()
        run = end_time - start_time
        run_time.append(run)
        if best_fitness - prev_fitness < epsilon: 
            sol = copy.deepcopy(best_sol)
            prev_fitness = best_fitness
    print(result)
    print(run_time)
    avg_step = [sum(step_avg)/ITE, sum(step_avg1)/ITE]
    wb = openpyxl.load_workbook('Book.xlsx')
    avg_time = sum(run_time)/len(run_time)
    sheet = wb['50']

    sheet.cell(row = k + 2, column = 1, value=dataList[k])
    for i, value in enumerate(result, start=1):
        sheet.cell(row= k + 2, column= 1 + i, value=value)
    sheet.cell(row = k + 2, column = 17, value = avg_time)
    for i, value in enumerate(avg_step, start=1):
        sheet.cell(row= k + 2, column= 18 + i, value=value)
    #print("đây nè", sol)
    sheet.cell(row = k + 2, column = 18, value = str(sol))
    wb.save('Book.xlsx')

    wb.close()
